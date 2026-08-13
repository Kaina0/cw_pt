"""
samples/ 配下にある測定結果(sample.xlsx と同じフォーマットの .xlsx)を
すべて読み込み、1つの集約帳票(output/aggregate_report.xlsx)にまとめる。

各シートのレイアウトは固定セル位置を前提としている(sample.xlsx 準拠):
    C3  : 測定日
    E3  : 測定者
    C6  : 製品番号
    C11 : 寸法測定値
    C12 : 寸法規格値 (例: "150±5")
    C13 : 重量測定値
    C14 : 重量規格値 (例: "1000±20")

集約帳票では、規格値("公称値±許容差")をパースして測定値が範囲内かどうかを判定し、
1行1サンプルの一覧表として出力する。

使い方:
    python3 aggregate_report.py [samples_dir] [output_path]
    (省略時は samples/ -> output/aggregate_report.xlsx)
"""

import datetime
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SAMPLES_DIR = BASE_DIR / "samples"
DEFAULT_OUTPUT_PATH = BASE_DIR / "output" / "aggregate_report.xlsx"

SPEC_PATTERN = re.compile(r"([+-]?\d+(?:\.\d+)?)\s*±\s*(\d+(?:\.\d+)?)")


def parse_spec(spec: object):
    """"150±5" のような規格値文字列を (下限, 上限) に変換する。パース不可なら None。"""
    if spec is None:
        return None
    match = SPEC_PATTERN.search(str(spec))
    if not match:
        return None
    nominal, tolerance = (float(match.group(1)), float(match.group(2)))
    return nominal - tolerance, nominal + tolerance


def judge(value: object, spec: object) -> str:
    if not isinstance(value, (int, float)):
        return "判定不能"
    bounds = parse_spec(spec)
    if bounds is None:
        return "判定不能"
    lower, upper = bounds
    return "合格" if lower <= value <= upper else "不合格"


def overall_judge(dim_judge: str, weight_judge: str) -> str:
    judges = {dim_judge, weight_judge}
    if "判定不能" in judges:
        return "判定不能"
    return "合格" if judges == {"合格"} else "不合格"


def load_measurement(path: Path) -> dict:
    """1件の帳票ファイルを読み込み、値と判定結果をまとめた dict を返す。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    dim_value = ws["C11"].value
    dim_spec = ws["C12"].value
    weight_value = ws["C13"].value
    weight_spec = ws["C14"].value
    dim_judge = judge(dim_value, dim_spec)
    weight_judge = judge(weight_value, weight_spec)

    return {
        "file_name": path.name,
        "measured_date": ws["C3"].value,
        "measurer": ws["E3"].value,
        "product_no": ws["C6"].value,
        "dim_value": dim_value,
        "dim_spec": dim_spec,
        "dim_judge": dim_judge,
        "weight_value": weight_value,
        "weight_spec": weight_spec,
        "weight_judge": weight_judge,
        "overall_judge": overall_judge(dim_judge, weight_judge),
    }


def collect_measurements(samples_dir: Path) -> list:
    paths = sorted(samples_dir.glob("*.xlsx"))
    if not paths:
        raise FileNotFoundError(f"サンプルファイルが見つかりません: {samples_dir}")
    return [load_measurement(p) for p in paths]


HEADER = [
    "ファイル名",
    "測定日",
    "測定者",
    "製品番号",
    "寸法測定値",
    "寸法規格値",
    "寸法判定",
    "重量測定値",
    "重量規格値",
    "重量判定",
    "総合判定",
]

# HEADER の各列に対応する measurement の dict キー
ROW_KEYS = [
    "file_name",
    "measured_date",
    "measurer",
    "product_no",
    "dim_value",
    "dim_spec",
    "dim_judge",
    "weight_value",
    "weight_spec",
    "weight_judge",
    "overall_judge",
]

THIN = Side(style="thin", color="000000")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
NG_FILL = PatternFill("solid", fgColor="FFC7CE")


def write_header(ws, row: int) -> None:
    for col, title in enumerate(HEADER, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def write_measurement_row(ws, row: int, measurement: dict) -> None:
    for col, key in enumerate(ROW_KEYS, start=1):
        value = measurement[key]
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        if col == 2 and isinstance(value, datetime.datetime):
            cell.number_format = "yyyy/mm/dd"
        if col in (7, 10, 11) and value == "不合格":
            cell.fill = NG_FILL


def write_summary(ws, row: int, measurements: list) -> None:
    total = len(measurements)
    ok_count = sum(1 for m in measurements if m["overall_judge"] == "合格")
    ng_count = sum(1 for m in measurements if m["overall_judge"] == "不合格")

    ws.cell(row=row, column=1, value="サンプル数").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)
    ws.cell(row=row + 1, column=1, value="合格").font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value=ok_count)
    ws.cell(row=row + 2, column=1, value="不合格").font = Font(bold=True)
    ws.cell(row=row + 2, column=2, value=ng_count)

    print(f"  合計 {total} 件 / 合格 {ok_count} 件 / 不合格 {ng_count} 件")


def set_column_widths(ws) -> None:
    widths = [22, 12, 10, 10, 12, 12, 10, 12, 12, 10, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_report(measurements: list, output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集約帳票"

    ws["A1"] = "測定結果 集約帳票"
    ws["A1"].font = Font(size=14, bold=True)

    header_row = 3
    write_header(ws, header_row)

    for r, measurement in enumerate(measurements, start=header_row + 1):
        write_measurement_row(ws, r, measurement)

    summary_row = header_row + len(measurements) + 2
    write_summary(ws, summary_row, measurements)

    set_column_widths(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"集約帳票を出力しました: {output_path}")


def main() -> None:
    samples_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SAMPLES_DIR
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT_PATH

    measurements = collect_measurements(samples_dir)
    build_report(measurements, output_path)


if __name__ == "__main__":
    main()
